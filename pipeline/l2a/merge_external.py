#!/usr/bin/env python3
"""
merge_external_ingredients.py
Extract ingredient/food names from all external data sources and merge into ingredient_seeds.json.
"""

import csv
import json
import os
import sys
from datetime import date
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[2]
EXTERNAL_DIR = BASE_DIR / "data" / "external"
OUTPUT_DIR = BASE_DIR / "output" / "l2a"
SEEDS_FILE = OUTPUT_DIR / "ingredient_seeds.json"
BY_SOURCE_FILE = OUTPUT_DIR / "ingredient_seeds_by_source.json"


def clean_item(item) -> str | None:
    if not item:
        return None
    item = str(item).strip()
    if len(item) < 2 or item.isdigit() or "http" in item.lower():
        return None
    return item


def extract_uk_cofid():
    results = []
    fpath = EXTERNAL_DIR / "uk_cofid" / "proximates.csv"
    try:
        with open(fpath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                item = clean_item(row.get("Food Name", ""))
                if item:
                    cat = (row.get("Food Group") or "").strip() or "other"
                    results.append((item, "uk_cofid", cat))
        print(f"  uk_cofid: {len(results)} items")
    except Exception as e:
        print(f"  WARNING uk_cofid: {e}")
    return results


def extract_japanese_mext():
    CAT_MAP = {
        "01": "grain", "02": "tuber", "03": "sugar", "04": "legume",
        "05": "nut", "06": "vegetable", "07": "fruit", "08": "mushroom",
        "09": "seaweed", "10": "fish", "11": "seafood", "12": "meat",
        "13": "egg", "14": "dairy", "15": "oil", "16": "confection",
        "17": "beverage", "18": "seasoning", "19": "convenience", "20": "other",
    }
    results = []
    fpath = EXTERNAL_DIR / "japanese_mext" / "general_composition_main.csv"
    try:
        with open(fpath, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)
            name_idx = 1  # 食品名
            code_idx = 0  # 食品番号
            for i, h in enumerate(header):
                if "食品名" in h:
                    name_idx = i
                if "食品番号" in h:
                    code_idx = i
            for row in reader:
                if len(row) <= name_idx:
                    continue
                item = clean_item(row[name_idx])
                if not item:
                    continue
                code = row[code_idx].strip() if len(row) > code_idx else ""
                cat = CAT_MAP.get(code[:2], "other")
                results.append((item, "japanese_mext", cat))
        print(f"  japanese_mext: {len(results)} items")
    except Exception as e:
        print(f"  WARNING japanese_mext: {e}")
    return results


def extract_korean_food_db():
    results = []
    fpath = EXTERNAL_DIR / "korean_food_db"
    try:
        csvs = list(fpath.glob("*.csv"))
        for cf in csvs:
            with open(cf, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    item = clean_item(row.get("식품명", ""))
                    if not item:
                        continue
                    cat = (row.get("식품대분류") or "").strip() or "other"
                    results.append((item, "korean_food_db", cat))
        print(f"  korean_food_db: {len(results)} items")
    except Exception as e:
        print(f"  WARNING korean_food_db: {e}")
    return results


def extract_chinese_recipes_kg():
    results = []
    fpath = EXTERNAL_DIR / "chinese_recipes_kg" / "meishichina_recipe.json"
    seen = set()
    try:
        with open(fpath, encoding="utf-8") as f:
            data = json.load(f)
        for recipe in data:
            for ing in recipe.get("ingredients", []):
                name = ing.get("name", "")
                item = clean_item(name)
                if item and item not in seen:
                    seen.add(item)
                    results.append((item, "chinese_recipes_kg", "other"))
        print(f"  chinese_recipes_kg: {len(results)} unique items")
    except Exception as e:
        print(f"  WARNING chinese_recipes_kg: {e}")
    return results


def extract_flavordb2():
    results = []
    fpath = EXTERNAL_DIR / "flavordb2" / "entity_flavor_molecules.csv"
    try:
        with open(fpath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                item = clean_item(row.get("entity_alias_readable", ""))
                if item:
                    cat = (row.get("entity_category") or "").strip() or "other"
                    results.append((item, "flavordb2", cat))
        print(f"  flavordb2: {len(results)} items")
    except Exception as e:
        print(f"  WARNING flavordb2: {e}")
    return results


def extract_aromadb():
    results = []
    fpath = EXTERNAL_DIR / "aromadb" / "VCF_supplement_data.xlsx"
    seen = set()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        header = None
        food_idx = food_cat_idx = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).lower().strip() if c else "" for c in row]
                for i, h in enumerate(header):
                    if "food name" in h:
                        food_idx = i
                    if "food category" in h:
                        food_cat_idx = i
                continue
            if food_idx is None:
                break
            val = row[food_idx] if food_idx < len(row) else None
            item = clean_item(val)
            if item and item not in seen:
                seen.add(item)
                cat = str(row[food_cat_idx]).strip() if food_cat_idx and food_cat_idx < len(row) and row[food_cat_idx] else "other"
                results.append((item, "aromadb", cat))
        wb.close()
        print(f"  aromadb: {len(results)} unique items")
    except Exception as e:
        print(f"  WARNING aromadb: {e}")
    return results


def extract_phenol_explorer():
    results = []
    fpath = EXTERNAL_DIR / "phenol_explorer" / "foods.csv"
    try:
        with open(fpath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                item = clean_item(row.get("name", ""))
                if item:
                    cat = (row.get("group") or "").strip() or "other"
                    results.append((item, "phenol_explorer", cat))
        print(f"  phenol_explorer: {len(results)} items")
    except Exception as e:
        print(f"  WARNING phenol_explorer: {e}")
    return results


def extract_foodmine():
    results = []
    fpath = EXTERNAL_DIR / "foodmine" / "FoodMine_GroceryProducts.xlsx"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        header = None
        name_idx = cat_idx = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).lower().strip() if c else "" for c in row]
                for i, h in enumerate(header):
                    if h == "name":
                        name_idx = i
                    if h == "category_name":
                        cat_idx = i
                continue
            if name_idx is None:
                break
            val = row[name_idx] if name_idx < len(row) else None
            item = clean_item(val)
            if item:
                cat = str(row[cat_idx]).strip() if cat_idx and cat_idx < len(row) and row[cat_idx] else "other"
                results.append((item, "foodmine", cat))
        wb.close()
        print(f"  foodmine: {len(results)} items")
    except Exception as e:
        print(f"  WARNING foodmine: {e}")
    return results


def extract_usda():
    """Extract unique food descriptions from USDA Foundation + SR Legacy."""
    results = []
    seen = set()
    usda_dir = EXTERNAL_DIR / "usda_fdc"

    # Find food.csv files (may be nested in subdirectories)
    food_csvs = list(usda_dir.rglob("food.csv"))
    if not food_csvs:
        print("  WARNING usda: no food.csv found")
        return results

    # Also load food_category if available
    cat_map = {}
    for cat_csv in usda_dir.rglob("food_category.csv"):
        try:
            with open(cat_csv, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    cid = row.get("id", "").strip().strip('"')
                    desc = row.get("description", "").strip().strip('"')
                    if cid and desc:
                        cat_map[cid] = desc
        except Exception:
            pass

    for food_csv in food_csvs:
        try:
            with open(food_csv, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    desc = (row.get("description") or "").strip().strip('"')
                    item = clean_item(desc)
                    if not item or item.lower() in seen:
                        continue
                    seen.add(item.lower())
                    cat_id = (row.get("food_category_id") or "").strip().strip('"')
                    cat = cat_map.get(cat_id, "other")
                    results.append((item, "usda_fdc", cat))
        except Exception as e:
            print(f"  WARNING usda ({food_csv.name}): {e}")

    print(f"  usda_fdc: {len(results)} unique items")
    return results


def main():
    print("=== Extracting from external sources ===")
    all_extracted = []
    all_extracted.extend(extract_uk_cofid())
    all_extracted.extend(extract_japanese_mext())
    all_extracted.extend(extract_korean_food_db())
    all_extracted.extend(extract_chinese_recipes_kg())
    all_extracted.extend(extract_flavordb2())
    all_extracted.extend(extract_aromadb())
    all_extracted.extend(extract_phenol_explorer())
    all_extracted.extend(extract_foodmine())
    all_extracted.extend(extract_usda())
    print(f"\nTotal extracted from external: {len(all_extracted)}")

    print("\n=== Loading existing ingredient_seeds.json ===")
    with open(SEEDS_FILE, encoding="utf-8") as f:
        seeds_data = json.load(f)

    # Handle both "items" and "ingredients" keys
    items_key = "items" if "items" in seeds_data else "ingredients"
    existing_items = seeds_data[items_key]
    print(f"Existing items: {len(existing_items)}")

    # Build lookup
    lookup = {}
    for idx, entry in enumerate(existing_items):
        key = entry["item"].lower()
        lookup[key] = idx
        if "sources" not in entry:
            entry["sources"] = ["books"]

    by_source = {"books": [e["item"] for e in existing_items]}

    new_count = 0
    for item, source, cat in all_extracted:
        key = item.lower()
        if key in lookup:
            entry = existing_items[lookup[key]]
            if source not in entry["sources"]:
                entry["sources"].append(source)
        else:
            new_entry = {
                "item": item,
                "frequency": 0,
                "category_guess": cat,
                "books": [],
                "sources": [source],
            }
            lookup[key] = len(existing_items)
            existing_items.append(new_entry)
            new_count += 1

        by_source.setdefault(source, [])
        if item not in by_source[source]:
            by_source[source].append(item)

    print(f"New items added: {new_count}")
    print(f"Total merged: {len(existing_items)}")

    # Update metadata
    meta = seeds_data.get("metadata", {})
    meta["version"] = "2.0"
    meta["total_items"] = len(existing_items)
    meta["merge_date"] = date.today().isoformat()
    meta["external_sources_merged"] = sorted(by_source.keys())
    seeds_data["metadata"] = meta

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(SEEDS_FILE, "w", encoding="utf-8") as f:
        json.dump(seeds_data, f, ensure_ascii=False, indent=2)
    print(f"\nWrote {SEEDS_FILE}")

    with open(BY_SOURCE_FILE, "w", encoding="utf-8") as f:
        json.dump(by_source, f, ensure_ascii=False, indent=2)
    print(f"Wrote {BY_SOURCE_FILE}")

    print("\n=== Source Summary ===")
    for src in sorted(by_source.keys()):
        print(f"  {src}: {len(by_source[src])} items")


if __name__ == "__main__":
    main()
